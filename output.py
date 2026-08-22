"""
DTS Pipeline — output.py
==========================
Generates the DTS daily Excel output.

Layout (matches production PDF):
  - One sheet per race, printed 1 race per page
  - Card summary sheet (Smart Value Plays + Longshot Lookers)
  - Per-race: header, column labels, horse rows with:
      • DTS Odds | Prob2Win | Runs (Early/Late/-)
      • Speed ★ | Jockey ★ | Trainer ★  (1-5 stars vs field)
      • Jockey / Trainer names
      • Smart Comment (betting angle headline)  — GREEN row when DTS < ML
      • Why-Like and Why-Fade attribution rows

Smart Comment logic is derived from the actual production pivot table which
uses three bucketed inputs:
  - DTS odds tier:      <1  |  1-2  |  2+
  - Morning Line tier:  <1  |  1-2  |  2+
  - Trainer ROI:        cold (<-.5) | soft (>-.5<0) | pos (>0<.25) | good (>.25<.5) | hot (>.5)
"""

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)

# ── Palette ──────────────────────────────────────────────────────────────────
C_NAVY        = "1F3864"
C_SLATE       = "2E4057"
C_WHITE       = "FFFFFF"
C_VALUE_BG    = "C6EFCE"   # green — DTS < ML
C_VALUE_FG    = "276221"
C_LONGSHOT_BG = "FFF2CC"   # gold  — longshot value
C_LONGSHOT_FG = "7D5A00"
C_COL_HDR     = "D9E1F2"   # blue-grey column headers
C_ALT         = "F5F5F5"   # alternating row
C_LIKE_BG     = "EBF5EB"
C_LIKE_FG     = "1A6B1A"
C_FADE_BG     = "FDEDED"
C_FADE_FG     = "8B0000"
C_SUMM_BG     = "E2EFDA"
C_RACE_TITLE  = "334975"   # race title bar


def _fill(color): return PatternFill("solid", fgColor=color)
def _font(bold=False, color="000000", size=9, italic=False, name="Arial"):
    return Font(bold=bold, color=color, size=size, italic=italic, name=name)
def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

_thin = Side(style="thin", color="CCCCCC")
_box  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


# ── Comment composer + value tier ──────────────────────────────────────────
# Replaces the legacy VLOOKUP pivot. The Comments column is now a two-
# sentence summary built from the attribution signals (the same ones that
# drive the ✓ Like / ✗ Fade panel) plus the DTS-vs-ML price comparison.
#
# Sentence 1: top attribute(s) by signal strength (|score|).
# Sentence 2: value vs morning line.
#
# A numeric `value_tier` (0-4) is also computed so downstream code (Best
# Bets card, best-bet row styling) can branch on price without grepping
# the prose. Tier mapping:
#   4 = Strong value      (DTS / ML  < 0.50)
#   3 = Real overlay      (0.50 - 0.75)
#   2 = Modest value      (0.75 - 0.90)
#   1 = Fairly priced     (0.90 - 1.15)
#   0 = Overbet or worse  (>= 1.15)

REASON_THRESHOLD = 0.20  # |score| above this counts as a "strong" signal


def _gather_signals(row) -> tuple[list, list]:
    """Pull (abs_score, text) tuples for like/fade signals from a row."""
    likes, fades = [], []
    for i in (1, 2, 3):
        for kind, bucket in (("like", likes), ("fade", fades)):
            text = row.get(f"why_{kind}_{i}", "")
            if text is None:
                continue
            try:
                if pd.isna(text):
                    continue
            except (TypeError, ValueError):
                pass
            text = str(text).strip()
            if not text or text.lower() == "nan":
                continue
            score = row.get(f"why_{kind}_{i}_score")
            try:
                s = abs(float(score)) if score is not None else 0.0
            except (TypeError, ValueError):
                s = 0.0
            bucket.append((s, text))
    likes.sort(reverse=True)
    fades.sort(reverse=True)
    return likes, fades


def _lc_first(s: str) -> str:
    """Lowercase the first char of a phrase for natural mid-sentence flow."""
    if not s:
        return s
    return s[0].lower() + s[1:]


def _value_tier(btsm, ml) -> int:
    """
    Map the model's win-probability EDGE over the scratch-adjusted morning line
    to a 0-4 value tier.

    EDGE = model win-prob / fair ML win-prob. `MornOddsAdj` (passed as `ml`) is
    renormalized to the model's own vig (VIG=1.2049 — the SAME vig baked into
    `DTSOdds`, passed as `btsm`), so both odds sit on the same overround and the
    vig cancels in the ratio. The de-vigged edge is therefore just:

        edge = (ml + 1) / (btsm + 1)      # = [1/(btsm+1)] / [1/(ml+1)]

    Bands are grounded in win-bet ROI at ~17% takeout, ROI ~= edge*(1-take) - 1:

        tier 4 : edge >= 2.00   very strong   (~+66% expected ROI)
        tier 3 : edge >= 1.50   GOLD-eligible (covers takeout + ~25% ROI target)
        tier 2 : edge >= 1.20   GREEN         (covers the ~17% takeout; +EV)
        tier 1 : edge >= 1.00   about right / no edge vs the line
        tier 0 : edge <  1.00   overbet (model shorter than the fair line)

    Shading gates downstream:
        GREEN = tier >= 2  (edge >= 1.20 - positive expected value)
        GOLD  = tier >= 3  (edge >= 1.50) AND top-half win-prob mass - see
                best_bet_flag(). Per-segment gates may override. Gold is
                intentionally rare; a race may have none.
    _value_phrase() still uses the underlying odds for wording nuance.
    """
    try:
        b = float(btsm); m = float(ml)
        if b < 0 or m < 0:
            return 0
        # de-vigged prob edge: model_prob / ml_prob = (ml+1)/(btsm+1).
        # Round to 6 places so exact boundary edges aren't lost to float error.
        edge = round((m + 1.0) / (b + 1.0), 6)
    except (TypeError, ValueError, ZeroDivisionError):
        return 0
    if edge >= 2.00: return 4
    if edge >= 1.50: return 3
    if edge >= 1.20: return 2
    if edge >= 1.00: return 1
    return 0


_CLAIMING_TYPES = {"C", "CO"}

def best_bet_flag(btsm, ml_adj, rank, field_size,
                  racetype=None, surface=None, track=None,
                  race_conditions=None, prob_above=None) -> bool:
    """
    GOLD best-bet gate. True only when value tier >= 3 (model win-prob >= 1.5x
    the fair, scratch-adjusted, de-vigged morning line - i.e. a 50% edge that
    covers the ~17% takeout plus a ~25% ROI target) AND a win-probability rank
    gate is met:
      - default             : cumulative win prob above the horse < 0.50
      - SAR turf claiming    : rank == 1 (top pick only)
      - SAR turf NY-bred     : cumulative win prob above the horse < 0.25

    The two "top X%" gates are measured on WIN-PROBABILITY MASS, not field
    position. prob_above is the summed ProbToWin of every horse ranked above
    this one, so a dominant favorite can push a positionally-top-half horse
    past the 50% mark and out of gold. Falls back to the positional rank/field
    test only when prob_above is not supplied (legacy callers / unit tests).

    Both tighter SAR-turf gates are evidence-based, measured on the SAR turf
    backtest (2006-2025, in-sample):

      * CLAIMING -> rank 1 only. Claiming gold bets below the top pick bleed
        (rank 2 -19%, rank 4 -25%) while the top pick returns +21% — the
        lower-ranked claiming "overlays" are longshots the model under-rates.

      * NY-BRED -> top 25%. NY-bred favorites are badly under-priced while the
        model's lower-half NY-bred picks lose: at the default top-50% gate
        NY-bred runs -5.5%, but tightening to top-25% flips it to +53% (rank-1
        +20%). Same favorite-bias signature as claiming, only stronger.

    A NY-bred RACE is identified by SAR + turf + the actual eligibility
    restriction text "FOALED IN NEW YORK" in RaceConditions1. A bare "NEW YORK"
    match (the original SAS rule) false-positives on open races that merely have
    "New York" in their name — e.g. the New York S. (a Grade I) and the New York
    Stallion Series — which would wrongly subject them to this tighter gate.
    Claiming takes precedence over the NY-bred gate (it is the stricter rule)
    for the rare NY-bred claiming turf race.

    Scoped to SAR + turf because that is where these were measured.
    Intentionally strict — many races will have zero gold best bets.
    """
    try:
        if _value_tier(btsm, ml_adj) < 3:
            return False
        # rank / field are needed only for the claiming gate and the positional
        # fallback; tolerate their absence so the prob_above path still works.
        try:
            r = float(rank); n = float(field_size)
        except (TypeError, ValueError):
            r = n = None
        is_sar_turf = (
            str(track).strip().upper() == "SAR"
            and str(surface).strip().upper() == "T"          # 'T'/'t' both -> turf
        )
        is_sar_turf_claim = (
            is_sar_turf
            and str(racetype).strip().upper() in _CLAIMING_TYPES
        )
        is_sar_turf_nybred = (
            is_sar_turf
            and "FOALED IN NEW YORK" in str(race_conditions).upper()  # NY-bred restricted race
        )
        if is_sar_turf_claim:
            return r is not None and r <= 1                  # SAR turf claiming: top pick only

        # The two "top X%" gates key on CUMULATIVE win probability — the share
        # of win prob held by horses ranked ABOVE this one (prob_above) — not
        # field position. A favorite-heavy race can push a positionally-top-half
        # horse past the 50% probability mark; that horse is not a best bet.
        if prob_above is not None:
            try:
                pa = float(prob_above)
            except (TypeError, ValueError):
                pa = None
            if pa is not None:
                return pa < (0.25 if is_sar_turf_nybred else 0.50)

        # Positional fallback (prob_above unavailable — legacy callers/tests).
        if r is None or n is None or n <= 0:
            return False
        if is_sar_turf_nybred:
            return (r / n) <= 0.25                           # SAR turf NY-bred: top 25%
        return (r / n) <= 0.5                                # everything else: top half
    except (TypeError, ValueError):
        return False


def green_flag(btsm, ml_adj, prob_above) -> bool:
    """
    GREEN 'longshot looker' flag (all tracks/models). True when a horse the
    model ranks in the BOTTOM half of the field by win-probability mass still
    shows a big edge over the fair, scratch-adjusted morning line: edge >= 1.75
    (model win-prob >= 1.75x the ML). These are live longshots to use
    underneath in exotics — NOT win bets (flat-win ROI is negative; the value
    is in the price and the tail). Deliberately limited (edge 1.75, ~30% of
    races) so GOLD stays the visual focus.

    Mutually exclusive with GOLD by construction: gold = top-half win-prob
    mass, green = bottom-half. A horse is at most one of the two.
    """
    try:
        pa = float(prob_above)
    except (TypeError, ValueError):
        return False
    if pa < 0.5:                        # top half -> gold territory, not green
        return False
    try:
        b = float(btsm); m = float(ml_adj)
        if b < 0 or m < 0:
            return False
        edge = (m + 1.0) / (b + 1.0)    # de-vigged prob edge (same basis as gold)
    except (TypeError, ValueError, ZeroDivisionError):
        return False
    return edge >= 1.75


def _value_phrase(tier: int, btsm, ml) -> str:
    """Render the value-vs-ML sentence, sensitive to absolute price."""
    try:
        b = float(btsm); m = float(ml)
    except (TypeError, ValueError):
        b = m = None

    is_short_fav = (b is not None and b < 2.5)

    if tier == 4:
        # DTS thinks horse is more than 2x as likely to win as ML
        if is_short_fav:
            return "Heavy favorite. DTS sees an even stronger chance than ML."
        return "Strong value vs ML, bet with conviction."
    if tier == 3:
        if is_short_fav:
            return "Short price, but DTS confirms the chalk."
        return "Real overlay at this price."
    if tier == 2:
        return "Modest value vs ML."
    if tier == 1:
        return "Priced about right."
    # tier 0: overbet or no data
    if b is not None and m is not None and m > 0 and b / m >= 1.5:
        return "Heavily overbet, pass at this price."
    if b is not None and m is not None:
        return "Overbet vs ML, needs a longer price."
    return ""


def compose_comment(row) -> str:
    """
    Two-sentence horse summary.

    Sentence 1 = top like/fade attributes by signal strength.
    Sentence 2 = value vs morning line.

    Requires row to have why_like_1..3, why_fade_1..3 with paired
    *_score columns, plus btsm_odds / ml_odds (or DTSOdds / MornOdds).
    """
    likes, fades = _gather_signals(row)
    strong_likes = [t for s, t in likes if s >= REASON_THRESHOLD]
    strong_fades = [t for s, t in fades if s >= REASON_THRESHOLD]

    # Build sentence 1
    if strong_likes and strong_fades:
        s1 = f"{strong_likes[0]}, but {_lc_first(strong_fades[0])}."
    elif len(strong_likes) >= 2:
        s1 = f"{strong_likes[0]} plus {_lc_first(strong_likes[1])}."
    elif strong_likes:
        s1 = f"{strong_likes[0]}."
    elif len(strong_fades) >= 2:
        s1 = f"{strong_fades[0]}; also {_lc_first(strong_fades[1])}."
    elif strong_fades:
        s1 = f"{strong_fades[0]}."
    else:
        s1 = "No standout attributes either way."

    # Build sentence 2 — value vs ML.
    # The TIER (gold/green/best-bet driver) is computed against the
    # SCRATCH-ADJUSTED morning line (MornOddsAdj) so it stays consistent with
    # the highlighting. The phrase wording only references absolute price,
    # never the adjusted number — nothing about the adjustment is surfaced.
    btsm = row.get("btsm_odds") if "btsm_odds" in row else row.get("DTSOdds")
    ml_raw = row.get("ml_odds") if "ml_odds" in row else row.get("MornOdds")
    ml_adj = row.get("MornOddsAdj")
    if ml_adj is None or (isinstance(ml_adj, float) and ml_adj != ml_adj):
        ml_adj = ml_raw
    tier = _value_tier(btsm, ml_adj)
    s2   = _value_phrase(tier, btsm, ml_adj)

    return f"{s1} {s2}".strip()


def value_tier(btsm, ml) -> int:
    """Public re-export for callers that need just the tier."""
    return _value_tier(btsm, ml)


# ── Star bars (1–5, race-relative) ─────────────────────────────────────────

def _stars(val, lo, hi, n=5) -> str:
    if pd.isna(val) or lo == hi:
        return "●●●○○"
    pct   = (float(val) - lo) / (hi - lo)
    count = max(1, min(n, round(1 + pct * (n - 1))))
    return "●" * count + "○" * (n - count)


def _runs(early_speed) -> str:
    if pd.isna(early_speed): return "  —  "
    v = float(early_speed)
    if v >  1.0: return "Early"
    if v < -1.0: return " Late"
    return "  —  "


def _fmt_odds(o) -> str:
    if pd.isna(o): return "—"
    o = float(o)
    return f"{o:.1f}"


def _title_case(s: str) -> str:
    return "/".join(p.strip().title() for p in str(s).split("/"))


# ── Main entry point ─────────────────────────────────────────────────────────

def generate_excel(
    scored_df: pd.DataFrame,
    feature_df: pd.DataFrame,
    output_path,
    config,
    track: str = "Keeneland",
    race_date: str = "",
    dirt_condition: str = "",
    turf_condition: str = "",
) -> Path:
    """
    Build DTS daily Excel — one race per sheet, 1 race per printed page.
    """
    output_path = Path(output_path)
    scored = scored_df[scored_df["ProbToWin"].notna()].copy()
    scored = scored.sort_values(["Race", "rank"])

    # Merge feature stats
    key = ["Track", "Date", "Race", "HorseName"]
    # Star-bar source columns (all on FIXED scales, see _build_race_sheet):
    #   SPD -> xBRISPd2     ((xBRISPd+14)^2,        fixed 1 to 729)
    #   JKY -> jckcm2_sarm  ((2.5+xJkyWCMstd)^2,    fixed 1 to 20.25)
    #   TRN -> trncm2_sart  ((2.5+trnwcm_sart)^2,   fixed 0.25 to 20.25)
    # PDF generation in run_pipeline.generate_pdf uses the same columns.
    feat_cols = ["xBRISPd2", "jckcm2_sarm",
                 "trncm2_sart", "EarlySpeed", "traineroutput"]
    feat_sub = feature_df[
        key + [c for c in feat_cols if c in feature_df.columns]
    ].copy()
    scored = scored.merge(feat_sub, on=key, how="left", suffixes=("", "_f"))

    # Compose the two-sentence Comments column and a numeric ValueTier
    # (0-4) derived from DTS-vs-ML. The legacy SmartComment/$$$ pivot is
    # gone; downstream logic should consult ValueTier (>= 3 == best bet).
    #
    # ValueTier + all value flagging below use the SCRATCH-ADJUSTED morning
    # line (MornOddsAdj, built in score._build_output). Displayed ML columns
    # still show the raw MornOdds. _ml_cmp is the behind-the-curtain Series
    # used for every DTS-vs-ML comparison in this file.
    if "MornOddsAdj" in scored.columns:
        _ml_cmp = scored["MornOddsAdj"].where(
            scored["MornOddsAdj"].notna(), scored["MornOdds"])
    else:
        _ml_cmp = scored["MornOdds"]
    scored["_MLCmp"] = _ml_cmp
    scored['Comments']  = scored.apply(compose_comment, axis=1)
    # GREEN tint  = green_flag (bottom-half win-prob + edge >= 1.75 longshot looker)
    # GOLD / best = BestBet flag (edge >= 1.5 AND top-half win-prob mass).
    #               Rare by design; a race may have none.
    scored['ValueTier'] = scored.apply(
        lambda r: value_tier(r.get('DTSOdds'), r.get('_MLCmp')), axis=1
    )
    # Per-race field size (post-scratch) — kept for the positional fallback.
    scored['_FieldSize'] = scored.groupby('Race')['Race'].transform('size')
    # Cumulative win probability held by horses ranked ABOVE each horse (the
    # summed ProbToWin of everything with a higher win prob in the race). This
    # drives the gold gate's "top X%" test on probability mass, not field
    # position — see best_bet_flag.
    _cum = (scored.sort_values(['Race', 'ProbToWin'], ascending=[True, False])
                  .groupby('Race')['ProbToWin'].cumsum())
    scored['_ProbAbove'] = _cum - scored['ProbToWin']
    scored['BestBet'] = scored.apply(
        lambda r: best_bet_flag(r.get('DTSOdds'), r.get('_MLCmp'),
                                r.get('rank'), r.get('_FieldSize'),
                                r.get('RaceType'), r.get('Surface'), r.get('Track'),
                                r.get('RaceConditions1'),
                                prob_above=r.get('_ProbAbove')), axis=1
    )
    scored['GreenFlag'] = scored.apply(
        lambda r: green_flag(r.get('DTSOdds'), r.get('_MLCmp'),
                             r.get('_ProbAbove')), axis=1
    )

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    races = sorted(scored["Race"].dropna().unique().astype(int))

    # ── Summary sheet ────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Card Summary")
    _build_summary(ws_sum, scored, track, race_date, dirt_condition, turf_condition)

    # ── One sheet per race ───────────────────────────────────────────────────
    for race_num in races:
        r_df = scored[scored["Race"] == race_num].copy()
        info = r_df.iloc[0]

        surf_map = {"D": "Dirt", "T": "Turf", "A": "All-Weather"}
        rt_map   = {
            "S": "Maiden SpWt", "M": "Maiden Claim",
            "C": "Claiming",    "A": "Allowance",
            "AO": "Allow Opt. Claim", "R": "Starter Alw",
            "G1": "Grade 1", "G2": "Grade 2", "G3": "Grade 3",
        }
        surf  = surf_map.get(str(info.Surface).upper(), str(info.Surface))
        rt    = rt_map.get(str(info.RaceType), str(info.RaceType))
        dist  = f"{info.Distanceinyards / 220:.1f}f"
        purse = f"${int(info.Purse/1000)}K" if pd.notna(info.Purse) else ""

        sheet_name = f"R{race_num} {rt[:12]}"
        ws = wb.create_sheet(sheet_name)
        _build_race_sheet(ws, r_df, race_num, rt, surf, dist, purse,
                          track, race_date, dirt_condition, turf_condition)

    wb.save(str(output_path))
    logger.info(f"Saved: {output_path}")
    return output_path


# ── Summary sheet ────────────────────────────────────────────────────────────

def _build_summary(ws, scored, track, race_date, dirt_condition, turf_condition):
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 35

    row = 1

    # Title
    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row, 1, f"{track}  ◆  {race_date}")
    c.font = _font(bold=True, size=14, color=C_WHITE)
    c.fill = _fill(C_NAVY)
    c.alignment = _align("center")
    ws.row_dimensions[row].height = 22
    row += 1

    ws.merge_cells(f"A{row}:F{row}")
    cond = f"Track Conditions:  Dirt: {dirt_condition or '—'}     Turf: {turf_condition or '—'}     No Scratches Updated"
    c = ws.cell(row, 1, cond)
    c.font = _font(size=9, color=C_WHITE, italic=True)
    c.fill = _fill(C_NAVY)
    c.alignment = _align("center")
    row += 2

    # Smart Value Plays header
    ws.merge_cells(f"A{row}:C{row}")
    c = ws.cell(row, 1, "Smart Value Plays of the Day")
    c.font = _font(bold=True, size=10, color=C_VALUE_FG)
    c.fill = _fill(C_SUMM_BG)
    row += 1

    # Best bets are the GOLD plays: BestBet flag (>=40% overlay AND win-prob
    # rank in the top half of the field). Rare by design. Comparison runs on
    # the scratch-adjusted line; display stays raw.
    big_bets = scored[scored["BestBet"]].sort_values("DTSOdds")

    for _, p in big_bets.head(6).iterrows():
        ws.merge_cells(f"A{row}:F{row}")
        txt = f"  Race {int(p.Race)}, #{int(p.Num)} - {p.HorseName} — {p.Comments}"
        c = ws.cell(row, 1, txt)
        c.font = _font(bold=True, italic=True, size=9, color=C_VALUE_FG)
        c.fill = _fill(C_SUMM_BG)
        row += 1

    row += 1

    # Longshot Lookers header
    ws.merge_cells(f"A{row}:C{row}")
    c = ws.cell(row, 1, "Longshot Lookers")
    c.font = _font(bold=True, size=10, color=C_LONGSHOT_FG)
    c.fill = _fill(C_LONGSHOT_BG)
    row += 1

    longshots = scored[
        (scored["DTSOdds"] >= 6) &
        (scored["DTSOdds"] < scored["_MLCmp"]) &
        (scored["rank"] <= 3)
    ].sort_values("DTSOdds").head(5)

    for _, p in longshots.iterrows():
        ws.merge_cells(f"A{row}:F{row}")
        txt = f"  Race {int(p.Race)}, #{int(p.Num)} - {p.HorseName}  (ML {_fmt_odds(p.MornOdds)})"
        c = ws.cell(row, 1, txt)
        c.font = _font(bold=True, italic=True, size=9, color=C_LONGSHOT_FG)
        c.fill = _fill(C_LONGSHOT_BG)
        row += 1

    row += 2

    # Full card table
    headers = ["#", "Horse", "ML", "DTS", "Win%", "Comments"]
    widths  = [5, 26, 8, 8, 8, 35]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
        c = ws.cell(row, ci, h)
        c.font      = _font(bold=True, size=9)
        c.fill      = _fill(C_COL_HDR)
        c.alignment = _align("center")
        c.border    = _box
    row += 1

    prev_race = None
    alt = False
    for _, horse in scored.iterrows():
        if horse.Race != prev_race:
            # Race divider
            ws.merge_cells(f"A{row}:F{row}")
            info = horse
            surf_map = {"D": "Dirt", "T": "Turf", "A": "All-Weather"}
            rt_map   = {
                "S":"Maiden SpWt","M":"Maiden Claim","C":"Claiming",
                "A":"Allowance","AO":"Allow Opt. Claim","R":"Starter Alw",
                "G1":"Grade 1","G2":"Grade 2","G3":"Grade 3",
            }
            surf = surf_map.get(str(info.Surface).upper(), "")
            rt   = rt_map.get(str(info.RaceType), str(info.RaceType))
            dist = f"{info.Distanceinyards/220:.1f}f"
            purse= f"${int(info.Purse/1000)}K" if pd.notna(info.Purse) else ""
            c = ws.cell(row, 1, f"  Race {int(horse.Race)}   {rt}   {surf}  —  {dist}  —  {purse}")
            c.font      = _font(bold=True, size=9, color=C_WHITE)
            c.fill      = _fill(C_SLATE)
            c.alignment = _align("left")
            ws.row_dimensions[row].height = 14
            row += 1
            prev_race = horse.Race
            alt = False

        is_val  = bool(horse.get("GreenFlag", False))               # green: bottom-half + edge>=1.75 longshot looker
        is_big  = bool(horse.get("BestBet", False))                  # gold: edge>=1.5 + top-half win-prob
        bg = C_LONGSHOT_BG if is_big else (C_VALUE_BG if is_val else (C_ALT if alt else C_WHITE))
        fg = C_VALUE_FG if is_val else "000000"

        row_data = [
            f"#{int(horse.Num)}",
            horse.HorseName,
            _fmt_odds(horse.MornOdds),
            _fmt_odds(horse.DTSOdds),
            f"{horse.ProbToWin:.0%}",
            horse.Comments,
        ]
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row, ci, val)
            c.fill      = _fill(bg)
            c.border    = _box
            c.font      = _font(
                bold   = (ci == 2 or (ci == 6 and is_big)),
                italic = (ci == 6 and is_big),
                color  = fg if (is_val and ci == 6) else (fg if is_val else "000000"),
                size   = 9,
            )
            c.alignment = _align("center" if ci in (1, 3, 4, 5) else "left",
                                  wrap=(ci == 6))
        ws.row_dimensions[row].height = 14
        row += 1
        alt = not alt

    # Print setup
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1


# ── Per-race sheet ────────────────────────────────────────────────────────────

# Column positions
_C = dict(
    NUM=1, HORSE=2, ML=3, DTS=4, PROB=5, RUNS=6,
    SPD=7, JKY=8, TRN=9, JT=10, COMMENT=11
)
_NCOLS = 11


def _build_race_sheet(ws, r_df, race_num, rt, surf, dist, purse,
                      track, race_date, dirt_condition, turf_condition):
    # Column widths
    widths = [4, 22, 5, 5, 6, 6, 7, 7, 7, 22, 32]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Star bar scales — all FIXED to the theoretical bounds of each
    # engineered transform. Lets a 4-star horse in Race 1 mean the same
    # caliber as a 4-star horse in Race 9.
    #   SPD: xBRISPd2     1.00 .. 729.00    ((-13+14)^2 .. (+13+14)^2)
    #   JKY: jckcm2_sarm  1.00 ..  20.25    ((2.5-1.5)^2 .. (2.5+2.0)^2)
    #   TRN: trncm2_sart  0.25 ..  20.25    ((2.5-2.0)^2 .. (2.5+2.0)^2)
    plo, phi = 1.0,  729.0
    jlo, jhi = 1.0,   20.25
    tlo, thi = 0.25,  20.25

    row = 1

    # ── Race header ──────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:K{row}")
    c = ws.cell(row, 1, f"{track}  ◆  {race_date}")
    c.font      = _font(bold=True, size=11, color=C_WHITE)
    c.fill      = _fill(C_NAVY)
    c.alignment = _align("center")
    ws.row_dimensions[row].height = 18
    row += 1

    ws.merge_cells(f"A{row}:K{row}")
    cond = f"Dirt: {dirt_condition or '—'}    Turf: {turf_condition or '—'}    No Scratches Updated"
    c = ws.cell(row, 1, cond)
    c.font      = _font(size=8, color=C_WHITE, italic=True)
    c.fill      = _fill(C_NAVY)
    c.alignment = _align("center")
    row += 1

    # Race title bar
    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row, 1, f"  Race {race_num}  {rt}")
    c.font      = _font(bold=True, size=10, color=C_WHITE)
    c.fill      = _fill(C_RACE_TITLE)
    c.alignment = _align("left")

    ws.merge_cells(f"G{row}:K{row}")
    c = ws.cell(row, 7, f"{surf}  —  {dist}  —  {purse}  ")
    c.font      = _font(bold=True, size=10, color=C_WHITE)
    c.fill      = _fill(C_RACE_TITLE)
    c.alignment = _align("right")
    ws.row_dimensions[row].height = 16
    row += 1

    # ── Column headers ────────────────────────────────────────────────────────
    col_labels = ["#", "Horse", "ML", "DTS\nOdds", "Prob\n2Win",
                  "Runs", "Speed", "Jockey", "Trainer",
                  "Jockey / Trainer\n(* jockey change)",
                  "Comments"]
    for ci, label in enumerate(col_labels, 1):
        c = ws.cell(row, ci, label)
        c.font      = _font(bold=True, size=8)
        c.fill      = _fill(C_COL_HDR)
        c.alignment = _align("center", wrap=True)
        c.border    = _box
    ws.row_dimensions[row].height = 24
    row += 1

    # ── Horses ────────────────────────────────────────────────────────────────
    alt = False
    for _, horse in r_df.iterrows():
        is_val   = bool(horse.get("GreenFlag", False))              # green: bottom-half + edge>=1.75 longshot looker
        is_big   = bool(horse.get("BestBet", False))                  # gold: edge>=1.5 + top-half win-prob
        bg       = C_LONGSHOT_BG if is_big else (C_VALUE_BG if is_val else (C_ALT if alt else C_WHITE))
        fg       = C_VALUE_FG if is_val else "000000"

        ppr_v = pd.to_numeric(horse.get("xBRISPd2"),    errors="coerce")
        jky_v = pd.to_numeric(horse.get("jckcm2_sarm"), errors="coerce")
        trn_v = pd.to_numeric(horse.get("trncm2_sart"), errors="coerce")
        es_v  = pd.to_numeric(horse.get("EarlySpeed"),            errors="coerce")

        jt_raw = f"{horse.TodaysJockey} / {horse.TodaysTrainer}"
        jt     = _title_case(jt_raw)

        comment   = horse.Comments
        ml_tag    = f"(ML{_fmt_odds(horse.MornOdds)})"
        full_cmt  = f"{comment}     {ml_tag}"

        row_vals = [
            f"#{int(horse.Num)}",
            horse.HorseName,
            _fmt_odds(horse.MornOdds),
            _fmt_odds(horse.DTSOdds),
            f"{horse.ProbToWin:.0%}",
            _runs(es_v),
            _stars(ppr_v, plo, phi),
            _stars(jky_v, jlo, jhi),
            _stars(trn_v, tlo, thi),
            jt,
            full_cmt,
        ]

        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row, ci, val)
            c.fill   = _fill(bg)
            c.border = _box
            c.font   = _font(
                bold   = (ci == _C["HORSE"] or (ci == _C["COMMENT"] and is_big)),
                italic = (ci == _C["COMMENT"] and is_big),
                color  = (C_VALUE_FG if is_val else "000000"),
                size   = 9,
            )
            c.alignment = _align(
                "center" if ci in (_C["NUM"], _C["ML"], _C["DTS"],
                                    _C["PROB"], _C["RUNS"],
                                    _C["SPD"], _C["JKY"], _C["TRN"]) else "left",
                wrap = (ci == _C["COMMENT"])
            )

        ws.row_dimensions[row].height = 15
        row += 1

        # ── Attribution rows ─────────────────────────────────────────────────
        likes = [horse.get(f"why_like_{i}", "") for i in range(1, 4)
                 if horse.get(f"why_like_{i}", "")]
        fades = [horse.get(f"why_fade_{i}", "") for i in range(1, 4)
                 if horse.get(f"why_fade_{i}", "")]

        if likes or fades:
            # blank cells A-B
            for ci in (1, 2):
                c = ws.cell(row, ci, "")
                c.fill = _fill("FFFFFF")

            # green likes block C-G
            ws.merge_cells(f"C{row}:G{row}")
            like_txt = "  ✅ " + "   |   ".join(likes) if likes else ""
            c = ws.cell(row, 3, like_txt)
            c.fill      = _fill(C_LIKE_BG)
            c.font      = _font(size=8, color=C_LIKE_FG, italic=True)
            c.alignment = _align("left", wrap=True)

            # red fades block H-K
            ws.merge_cells(f"H{row}:K{row}")
            fade_txt = "  ❌ " + "   |   ".join(fades) if fades else ""
            c = ws.cell(row, 8, fade_txt)
            c.fill      = _fill(C_FADE_BG)
            c.font      = _font(size=8, color=C_FADE_FG, italic=True)
            c.alignment = _align("left", wrap=True)

            ws.row_dimensions[row].height = 26
            row += 1

        alt = not alt

    # ── Spacing + footer ──────────────────────────────────────────────────────
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row, 1, "Daily $10    ·    Full Meet $75")
    c.font      = _font(size=8, color="888888")
    c.alignment = _align("left")

    ws.merge_cells(f"G{row}:K{row}")
    c = ws.cell(row, 7, "www.BeTheSmartMoney.com")
    c.font      = _font(size=8, color="888888", italic=True)
    c.alignment = _align("right")

    # ── Print settings ────────────────────────────────────────────────────────
    ws.page_setup.orientation       = "landscape"
    ws.page_setup.paperSize         = 9   # A4; use 1 for Letter
    ws.page_setup.fitToPage         = True
    ws.page_setup.fitToWidth        = 1
    ws.page_setup.fitToHeight       = 1   # force 1 page tall
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # Print area covers all used rows
    ws.print_area = f"A1:{get_column_letter(_NCOLS)}{row}"
    ws.freeze_panes = "A5"
