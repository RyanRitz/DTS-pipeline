"""
BTSM Pipeline — output.py
==========================
Generates the BTSM daily Excel output.

Layout (matches production PDF):
  - One sheet per race, printed 1 race per page
  - Card summary sheet (Smart Value Plays + Longshot Lookers)
  - Per-race: header, column labels, horse rows with:
      • BTSM Odds | Prob2Win | Runs (Early/Late/-)
      • Speed ★ | Jockey ★ | Trainer ★  (1-5 stars vs field)
      • Jockey / Trainer names
      • Smart Comment (betting angle headline)  — GREEN row when BTSM < ML
      • Why-Like and Why-Fade attribution rows

Smart Comment logic is derived from the actual production pivot table which
uses three bucketed inputs:
  - BTSM odds tier:     <1  |  1-2  |  2+
  - Morning Line tier:  <1  |  1-2  |  2+
  - Trainer ROI:        cold (<-.5) | soft (>-.5<0) | pos (>0<.25) | good (>.25<.5) | hot (>.5)
"""

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)

# ── Palette ──────────────────────────────────────────────────────────────────
C_NAVY        = "1F3864"
C_SLATE       = "2E4057"
C_WHITE       = "FFFFFF"
C_VALUE_BG    = "C6EFCE"   # green — BTSM < ML
C_VALUE_FG    = "276221"
C_LONGSHOT_BG = "FFF2CC"   # gold  — longshot value
C_LONGSHOT_FG = "7D5A00"
C_COL_HDR     = "D9E1F2"   # blue-grey column headers
C_ALT         = "F5F5F5"   # alternating row
C_LIKE_BG     = "EBF5EB"
C_LIKE_FG     = "1A6B1A"
C_FADE_BG     = "FDEDED"
C_FADE_FG     = "8B0000"
C_SUMM_BG     = "E2EFDA"
C_RACE_TITLE  = "334975"   # race title bar


def _fill(color): return PatternFill("solid", fgColor=color)
def _font(bold=False, color="000000", size=9, italic=False, name="Arial"):
    return Font(bold=bold, color=color, size=size, italic=italic, name=name)
def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

_thin = Side(style="thin", color="CCCCCC")
_box  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


# ── Comment logic from pivot matrix ─────────────────────────────────────────

# ── Comment VLOOKUP table — exact replica of Excel pivot BL3:BM48 ──────────
_VLOOKUP = {
    "<1<1<-.5":       "Not a strong contender",
    "1-2<1<-.5":      "Has Shot; bad price",
    "2+<1<-.5":       "Good horse; bad price",
    "<11-2<-.5":      "A crowd favorite, no value",
    "1-21-2<-.5":     "Decent horse; not much value",
    "2+1-2<-.5":      "Solid horse; not much value",
    "<12+<-.5":       "A crowd favorite, no value",
    "1-22+<-.5":      "Decent horse; no value",
    "2+2+<-.5":       "One to beat; bet if live value",
    "<1<1>-.5<0":     "Long shot - no value",
    "1-2<1>-.5<0":    "Good horse; possible value",
    "2+<1>-.5<0":     "Decent bet; possible price",
    "<11-2>-.5<0":    "Solid horse; limited value",
    "1-21-2>-.5<0":   "Good shot; need a price",
    "2+1-2>-.5<0":    "Great shot; need a price",
    "<12+>-.5<0":     "A lot to like; BTSM doesn't",
    "1-22+>-.5<0":    "A lot to like; will be overbet",
    "2+2+>-.5<0":     "One to beat; bet if live value",
    "<1<1>0<.25":     "Long shot - worth a look",
    "1-2<1>0<.25":    "Long shot - worthy of wager",
    "2+<1>0<.25":     "Long shot - really like",
    "<11-2>0<.25":    "SHOULDN'T HAPPEN",
    "1-21-2>0<.25":   "Good horse may show value",
    "2+1-2>0<.25":    "Great horse should show value",
    "<12+>0<.25":     "SHOULDN'T HAPPEN",
    "1-22+>0<.25":    "SHOULDN'T HAPPEN",
    "2+2+>0<.25":     "Great chance; worth a look",
    "<1<1>.25<.5":    "BTSM Longshot ValueBet $",
    "1-2<1>.25<.5":   "BTSM Longshot ValueBet $",
    "2+<1>.25<.5":    "BTSM Longshot ValueBet $$",
    "<11-2>.25<.5":   "SHOULDN'T HAPPEN",
    "1-21-2>.25<.5":  "BTSM Best ValueBet $",
    "2+1-2>.25<.5":   "BTSM Best ValueBet $$",
    "<12+>.25<.5":    "SHOULDN'T HAPPEN",
    "1-22+>.25<.5":   "SHOULDN'T HAPPEN",
    "2+2+>.25<.5":    "BTSM Best ValueBet $$$",
    "<1<1>.5":        "BTSM Longshot ValueBet $",
    "1-2<1>.5":       "BTSM Longshot ValueBet $",
    "2+<1>.5":        "BTSM Longshot ValueBet $",
    "<11-2>.5":       "SHOULDN'T HAPPEN",
    "1-21-2>.5":      "BTSM Best ValueBet $$",
    "2+1-2>.5":       "BTSM Best ValueBet $$$",
    "<12+>.5":        "SHOULDN'T HAPPEN",
    "1-22+>.5":       "SHOULDN'T HAPPEN",
    "2+2+>.5":        "BTSM Best ValueBet $$$$",
}


def _t3(v):
    """Three-tier BTSM/ML FairShare category."""
    if v >= 2: return "2+"
    if v >= 1: return "1-2"
    return "<1"


def _roi_t(v):
    """Five-tier ROI category."""
    if v >= 0.5:   return ">.5"
    if v >= 0.25:  return ">.25<.5"
    if v >= 0:     return ">0<.25"
    if v > -0.5:   return ">-.5<0"
    return "<-.5"


def smart_comment(prob: float, ml: float, n: int, af: float) -> str:
    """
    Exact replication of the Excel pivot comment formula.

    Parameters
    ----------
    prob : ProbToWin (normalised win probability)
    ml   : MornOdds (morning line odds)
    n    : NumOfEntries (horses in the race)
    af   : Race-level sum of ML implied probs = sum(1/(ML+1)) for all horses

    Logic (mirrors Excel columns BB, BC, BF → BH → VLOOKUP):
      BB = prob × n                           (BTSM FairShare expected wins)
      AD = 1/(ml+1)                           (ML implied probability)
      AG = AD / af                            (ML implied prob normalised)
      AH = AG × 1.2
      AI = 1/AH − 1                          (fair-value ML odds)
      BC = ((1/(AI+1))/1.28) × n             (ML FairShare)
      AL = (prob × AI)/(1−prob) − 1          (horse-level ROI)
      key = tier(BB) + tier(BC) + roi_tier(AL)
    """
    if pd.isna(prob) or pd.isna(ml) or pd.isna(n) or not af:
        return "Long shot - no value"
    prob, ml, n, af = float(prob), float(ml), int(n), float(af)
    bb  = prob * n
    ad  = 1.0 / (ml + 1)
    ag  = ad / af
    ah  = ag * 1.2
    ai  = (1.0 / ah) - 1 if ah else 0
    bc  = ((1.0 / (ai + 1)) / 1.28) * n if ai > -1 else 0
    al  = (prob * ai) / (1 - prob) - 1 if prob < 1 else 0
    key = _t3(bb) + _t3(bc) + _roi_t(al)
    return _VLOOKUP.get(key, "Long shot - no value")


# ── Star bars (1–5, race-relative) ─────────────────────────────────────────

def _stars(val, lo, hi, n=5) -> str:
    if pd.isna(val) or lo == hi:
        return "●●●○○"
    pct   = (float(val) - lo) / (hi - lo)
    count = max(1, min(n, round(1 + pct * (n - 1))))
    return "●" * count + "○" * (n - count)


def _runs(early_speed) -> str:
    if pd.isna(early_speed): return "  —  "
    v = float(early_speed)
    if v >  1.0: return "Early"
    if v < -1.0: return " Late"
    return "  —  "


def _fmt_odds(o) -> str:
    if pd.isna(o): return "—"
    o = float(o)
    return f"{o:.1f}"


def _title_case(s: str) -> str:
    return "/".join(p.strip().title() for p in str(s).split("/"))


# ── Main entry point ─────────────────────────────────────────────────────────

def generate_excel(
    scored_df: pd.DataFrame,
    feature_df: pd.DataFrame,
    output_path,
    config,
    track: str = "Keeneland",
    race_date: str = "",
    dirt_condition: str = "",
    turf_condition: str = "",
) -> Path:
    """
    Build BTSM daily Excel — one race per sheet, 1 race per printed page.
    """
    output_path = Path(output_path)
    scored = scored_df[scored_df["ProbToWin"].notna()].copy()
    scored = scored.sort_values(["Race", "rank"])

    # Merge feature stats
    key = ["Track", "Date", "Race", "HorseName"]
    feat_cols = ["BRISPrimePowerRating", "JockeyCurMtWPpct",
                 "TrainerCurMtWPpct", "EarlySpeed", "traineroutput"]
    feat_sub = feature_df[
        key + [c for c in feat_cols if c in feature_df.columns]
    ].copy()
    scored = scored.merge(feat_sub, on=key, how="left", suffixes=("", "_f"))

    # Compute AF = sum(1/(ML+1)) per race — mirrors Excel pivot AF column
    scored['_ml_implied'] = 1.0 / (scored['MornOdds'] + 1)
    scored['_af']         = scored.groupby('Race')['_ml_implied'].transform('sum')

    # Add Smart Comment to every horse using the exact pivot formula
    scored['SmartComment'] = scored.apply(
        lambda r: smart_comment(r.ProbToWin, r.MornOdds,
                                r.NumOfEntries, r._af), axis=1
    )

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    races = sorted(scored["Race"].dropna().unique().astype(int))

    # ── Summary sheet ────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Card Summary")
    _build_summary(ws_sum, scored, track, race_date, dirt_condition, turf_condition)

    # ── One sheet per race ───────────────────────────────────────────────────
    for race_num in races:
        r_df = scored[scored["Race"] == race_num].copy()
        info = r_df.iloc[0]

        surf_map = {"D": "Dirt", "T": "Turf", "A": "All-Weather"}
        rt_map   = {
            "S": "Maiden SpWt", "M": "Maiden Claim",
            "C": "Claiming",    "A": "Allowance",
            "AO": "Allow Opt. Claim", "R": "Starter Alw",
            "G1": "Grade 1", "G2": "Grade 2", "G3": "Grade 3",
        }
        surf  = surf_map.get(str(info.Surface).upper(), str(info.Surface))
        rt    = rt_map.get(str(info.RaceType), str(info.RaceType))
        dist  = f"{info.Distanceinyards / 220:.1f}f"
        purse = f"${int(info.Purse/1000)}K" if pd.notna(info.Purse) else ""

        sheet_name = f"R{race_num} {rt[:12]}"
        ws = wb.create_sheet(sheet_name)
        _build_race_sheet(ws, r_df, race_num, rt, surf, dist, purse,
                          track, race_date, dirt_condition, turf_condition)

    wb.save(str(output_path))
    logger.info(f"Saved: {output_path}")
    return output_path


# ── Summary sheet ────────────────────────────────────────────────────────────

def _build_summary(ws, scored, track, race_date, dirt_condition, turf_condition):
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 35

    row = 1

    # Title
    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row, 1, f"{track}  ◆  {race_date}")
    c.font = _font(bold=True, size=14, color=C_WHITE)
    c.fill = _fill(C_NAVY)
    c.alignment = _align("center")
    ws.row_dimensions[row].height = 22
    row += 1

    ws.merge_cells(f"A{row}:F{row}")
    cond = f"Track Conditions:  Dirt: {dirt_condition or '—'}     Turf: {turf_condition or '—'}     No Scratches Updated"
    c = ws.cell(row, 1, cond)
    c.font = _font(size=9, color=C_WHITE, italic=True)
    c.fill = _fill(C_NAVY)
    c.alignment = _align("center")
    row += 2

    # Smart Value Plays header
    ws.merge_cells(f"A{row}:C{row}")
    c = ws.cell(row, 1, "Smart Value Plays of the Day")
    c.font = _font(bold=True, size=10, color=C_VALUE_FG)
    c.fill = _fill(C_SUMM_BG)
    row += 1

    value_plays = scored[scored["BTSMOdds"] < scored["MornOdds"]].sort_values("BTSMOdds")
    big_bets    = value_plays[scored["SmartComment"].str.contains(r"\$\$", na=False)]

    for _, p in big_bets.head(6).iterrows():
        ws.merge_cells(f"A{row}:F{row}")
        txt = f"  Race {int(p.Race)}, #{int(p.Num)} - {p.HorseName} is a {p.SmartComment}"
        c = ws.cell(row, 1, txt)
        c.font = _font(bold=True, italic=True, size=9, color=C_VALUE_FG)
        c.fill = _fill(C_SUMM_BG)
        row += 1

    row += 1

    # Longshot Lookers header
    ws.merge_cells(f"A{row}:C{row}")
    c = ws.cell(row, 1, "Longshot Lookers")
    c.font = _font(bold=True, size=10, color=C_LONGSHOT_FG)
    c.fill = _fill(C_LONGSHOT_BG)
    row += 1

    longshots = scored[
        (scored["BTSMOdds"] >= 6) &
        (scored["BTSMOdds"] < scored["MornOdds"]) &
        (scored["rank"] <= 3)
    ].sort_values("BTSMOdds").head(5)

    for _, p in longshots.iterrows():
        ws.merge_cells(f"A{row}:F{row}")
        txt = f"  Race {int(p.Race)}, #{int(p.Num)} - {p.HorseName}  (ML {_fmt_odds(p.MornOdds)})"
        c = ws.cell(row, 1, txt)
        c.font = _font(bold=True, italic=True, size=9, color=C_LONGSHOT_FG)
        c.fill = _fill(C_LONGSHOT_BG)
        row += 1

    row += 2

    # Full card table
    headers = ["#", "Horse", "ML", "BTSM", "Win%", "Smart Comment"]
    widths  = [5, 26, 8, 8, 8, 35]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
        c = ws.cell(row, ci, h)
        c.font      = _font(bold=True, size=9)
        c.fill      = _fill(C_COL_HDR)
        c.alignment = _align("center")
        c.border    = _box
    row += 1

    prev_race = None
    alt = False
    for _, horse in scored.iterrows():
        if horse.Race != prev_race:
            # Race divider
            ws.merge_cells(f"A{row}:F{row}")
            info = horse
            surf_map = {"D": "Dirt", "T": "Turf", "A": "All-Weather"}
            rt_map   = {
                "S":"Maiden SpWt","M":"Maiden Claim","C":"Claiming",
                "A":"Allowance","AO":"Allow Opt. Claim","R":"Starter Alw",
                "G1":"Grade 1","G2":"Grade 2","G3":"Grade 3",
            }
            surf = surf_map.get(str(info.Surface).upper(), "")
            rt   = rt_map.get(str(info.RaceType), str(info.RaceType))
            dist = f"{info.Distanceinyards/220:.1f}f"
            purse= f"${int(info.Purse/1000)}K" if pd.notna(info.Purse) else ""
            c = ws.cell(row, 1, f"  Race {int(horse.Race)}   {rt}   {surf}  —  {dist}  —  {purse}")
            c.font      = _font(bold=True, size=9, color=C_WHITE)
            c.fill      = _fill(C_SLATE)
            c.alignment = _align("left")
            ws.row_dimensions[row].height = 14
            row += 1
            prev_race = horse.Race
            alt = False

        is_val  = bool(horse.BTSMOdds < horse.MornOdds)
        is_big  = "$$" in str(horse.SmartComment)
        bg = C_VALUE_BG if is_val else (C_ALT if alt else C_WHITE)
        fg = C_VALUE_FG if is_val else "000000"

        row_data = [
            f"#{int(horse.Num)}",
            horse.HorseName,
            _fmt_odds(horse.MornOdds),
            _fmt_odds(horse.BTSMOdds),
            f"{horse.ProbToWin:.0%}",
            horse.SmartComment,
        ]
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row, ci, val)
            c.fill      = _fill(bg)
            c.border    = _box
            c.font      = _font(
                bold   = (ci == 2 or (ci == 6 and is_big)),
                italic = (ci == 6 and is_big),
                color  = fg if (is_val and ci == 6) else (fg if is_val else "000000"),
                size   = 9,
            )
            c.alignment = _align("center" if ci in (1, 3, 4, 5) else "left",
                                  wrap=(ci == 6))
        ws.row_dimensions[row].height = 14
        row += 1
        alt = not alt

    # Print setup
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1


# ── Per-race sheet ────────────────────────────────────────────────────────────

# Column positions
_C = dict(
    NUM=1, HORSE=2, ML=3, BTSM=4, PROB=5, RUNS=6,
    SPD=7, JKY=8, TRN=9, JT=10, COMMENT=11
)
_NCOLS = 11


def _build_race_sheet(ws, r_df, race_num, rt, surf, dist, purse,
                      track, race_date, dirt_condition, turf_condition):
    # Column widths
    widths = [4, 22, 5, 5, 6, 6, 7, 7, 7, 22, 32]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Race-level stat ranges for star bars
    pprs = pd.to_numeric(r_df.get("BRISPrimePowerRating", pd.Series()), errors="coerce")
    jkys = pd.to_numeric(r_df.get("JockeyCurMtWPpct",     pd.Series()), errors="coerce")
    trns = pd.to_numeric(r_df.get("TrainerCurMtWPpct",    pd.Series()), errors="coerce")
    plo, phi = pprs.min(), pprs.max()
    jlo, jhi = jkys.min(), jkys.max()
    tlo, thi = trns.min(), trns.max()

    row = 1

    # ── Race header ──────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:K{row}")
    c = ws.cell(row, 1, f"{track}  ◆  {race_date}")
    c.font      = _font(bold=True, size=11, color=C_WHITE)
    c.fill      = _fill(C_NAVY)
    c.alignment = _align("center")
    ws.row_dimensions[row].height = 18
    row += 1

    ws.merge_cells(f"A{row}:K{row}")
    cond = f"Dirt: {dirt_condition or '—'}    Turf: {turf_condition or '—'}    No Scratches Updated"
    c = ws.cell(row, 1, cond)
    c.font      = _font(size=8, color=C_WHITE, italic=True)
    c.fill      = _fill(C_NAVY)
    c.alignment = _align("center")
    row += 1

    # Race title bar
    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row, 1, f"  Race {race_num}  {rt}")
    c.font      = _font(bold=True, size=10, color=C_WHITE)
    c.fill      = _fill(C_RACE_TITLE)
    c.alignment = _align("left")

    ws.merge_cells(f"G{row}:K{row}")
    c = ws.cell(row, 7, f"{surf}  —  {dist}  —  {purse}  ")
    c.font      = _font(bold=True, size=10, color=C_WHITE)
    c.fill      = _fill(C_RACE_TITLE)
    c.alignment = _align("right")
    ws.row_dimensions[row].height = 16
    row += 1

    # ── Column headers ────────────────────────────────────────────────────────
    col_labels = ["#", "Horse", "ML", "BTSM\nOdds", "Prob\n2Win",
                  "Runs", "Speed", "Jockey", "Trainer",
                  "Jockey / Trainer\n(* jockey change)",
                  "Smart Comments\n($=sml; $$=med; $$$=big bet)"]
    for ci, label in enumerate(col_labels, 1):
        c = ws.cell(row, ci, label)
        c.font      = _font(bold=True, size=8)
        c.fill      = _fill(C_COL_HDR)
        c.alignment = _align("center", wrap=True)
        c.border    = _box
    ws.row_dimensions[row].height = 24
    row += 1

    # ── Horses ────────────────────────────────────────────────────────────────
    alt = False
    for _, horse in r_df.iterrows():
        is_val   = bool(horse.BTSMOdds < horse.MornOdds)
        is_big   = "$$" in str(horse.SmartComment)
        bg       = C_VALUE_BG if is_val else (C_ALT if alt else C_WHITE)
        fg       = C_VALUE_FG if is_val else "000000"

        ppr_v = pd.to_numeric(horse.get("BRISPrimePowerRating"), errors="coerce")
        jky_v = pd.to_numeric(horse.get("JockeyCurMtWPpct"),     errors="coerce")
        trn_v = pd.to_numeric(horse.get("TrainerCurMtWPpct"),    errors="coerce")
        es_v  = pd.to_numeric(horse.get("EarlySpeed"),            errors="coerce")

        jt_raw = f"{horse.TodaysJockey} / {horse.TodaysTrainer}"
        jt     = _title_case(jt_raw)

        comment   = horse.SmartComment
        ml_tag    = f"(ML{_fmt_odds(horse.MornOdds)})"
        full_cmt  = f"{comment}     {ml_tag}"

        row_vals = [
            f"#{int(horse.Num)}",
            horse.HorseName,
            _fmt_odds(horse.MornOdds),
            _fmt_odds(horse.BTSMOdds),
            f"{horse.ProbToWin:.0%}",
            _runs(es_v),
            _stars(ppr_v, plo, phi),
            _stars(jky_v, jlo, jhi),
            _stars(trn_v, tlo, thi),
            jt,
            full_cmt,
        ]

        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row, ci, val)
            c.fill   = _fill(bg)
            c.border = _box
            c.font   = _font(
                bold   = (ci == _C["HORSE"] or (ci == _C["COMMENT"] and is_big)),
                italic = (ci == _C["COMMENT"] and is_big),
                color  = (C_VALUE_FG if is_val else "000000"),
                size   = 9,
            )
            c.alignment = _align(
                "center" if ci in (_C["NUM"], _C["ML"], _C["BTSM"],
                                    _C["PROB"], _C["RUNS"],
                                    _C["SPD"], _C["JKY"], _C["TRN"]) else "left",
                wrap = (ci == _C["COMMENT"])
            )

        ws.row_dimensions[row].height = 15
        row += 1

        # ── Attribution rows ─────────────────────────────────────────────────
        likes = [horse.get(f"why_like_{i}", "") for i in range(1, 4)
                 if horse.get(f"why_like_{i}", "")]
        fades = [horse.get(f"why_fade_{i}", "") for i in range(1, 4)
                 if horse.get(f"why_fade_{i}", "")]

        if likes or fades:
            # blank cells A-B
            for ci in (1, 2):
                c = ws.cell(row, ci, "")
                c.fill = _fill("FFFFFF")

            # green likes block C-G
            ws.merge_cells(f"C{row}:G{row}")
            like_txt = "  ✅ " + "   |   ".join(likes) if likes else ""
            c = ws.cell(row, 3, like_txt)
            c.fill      = _fill(C_LIKE_BG)
            c.font      = _font(size=8, color=C_LIKE_FG, italic=True)
            c.alignment = _align("left", wrap=True)

            # red fades block H-K
            ws.merge_cells(f"H{row}:K{row}")
            fade_txt = "  ❌ " + "   |   ".join(fades) if fades else ""
            c = ws.cell(row, 8, fade_txt)
            c.fill      = _fill(C_FADE_BG)
            c.font      = _font(size=8, color=C_FADE_FG, italic=True)
            c.alignment = _align("left", wrap=True)

            ws.row_dimensions[row].height = 26
            row += 1

        alt = not alt

    # ── Spacing + footer ──────────────────────────────────────────────────────
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row, 1, "Daily $10    ·    Full Meet $75")
    c.font      = _font(size=8, color="888888")
    c.alignment = _align("left")

    ws.merge_cells(f"G{row}:K{row}")
    c = ws.cell(row, 7, "www.BeTheSmartMoney.com")
    c.font      = _font(size=8, color="888888", italic=True)
    c.alignment = _align("right")

    # ── Print settings ────────────────────────────────────────────────────────
    ws.page_setup.orientation       = "landscape"
    ws.page_setup.paperSize         = 9   # A4; use 1 for Letter
    ws.page_setup.fitToPage         = True
    ws.page_setup.fitToWidth        = 1
    ws.page_setup.fitToHeight       = 1   # force 1 page tall
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # Print area covers all used rows
    ws.print_area = f"A1:{get_column_letter(_NCOLS)}{row}"
    ws.freeze_panes = "A5"
